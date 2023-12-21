import urllib

import dateutil
import pandas as pd
import os
import openpyxl
from dotenv import load_dotenv
from fuzzywuzzy import process
from sqlalchemy import create_engine, text
from sqlalchemy.engine import cursor

from datetime import datetime
from et_xmlfile import xmlfile
from openpyxl import load_workbook
import numpy
import dateutil
import pytz
import six
import tzdata


servers = [
    {
        'server_name': "q3solutions-dev.database.windows.net",
        'database_name': "PropertyManager-dev",
        'username': "q3solutions-dev",
        'password': "Q3testMcQ"
    },
    """{
        'server_name': "q3solutions.database.windows.net",
        'database_name': "PropertyManager",
        'username': "burtonguster",
        'password': "heardaboutplut0!"
    },
    {
        'server_name': "q3baldwinfi.database.windows.net",
        'database_name': "PropertyManager",
        'username': "baldwinfi",
        'password': "GObucks2nite!"
    },
    {
        'server_name': "q3aberdeenlegacy.database.windows.net",
        'database_name': "PropertyManager",
        'username': "CloudSAd02d20b4",
        'password': "C0mm1t4ppr0v3d!"
    }"""
]
# Set up the Excel file path
# file_path = r'C:\Users\rdsku\Aberdeen Capital\Aberdeen Capital - Documents\Underwriting Templates\GreenIQ\RentRoll.xlsx'
#file_path = "Cirrus_Rent_Roll_with_Lease_Charges_40501.xlsx"
file_path = "C:/xampp/htdocs/RentInsite/Q3/orchardPark/ceva_americana-main/Rent_Roll_40501.xlsx"

# sheet_name = 'RentRoll'

def read_file(file_name):
    df = pd.read_excel(file_name,
            sheet_name = 0)
    global asofdate
    asofdate = df.iloc[5,0].split()[-1]

    index = df[df.iloc[:, 0] == 'Unit'].index[0]
    df = pd.read_excel(file_name,
            sheet_name = 0,
            skiprows=index+1)
    return df

def build_headers(df):
    # Since the headers are split in two rows E.G. Unit + Sq Ft, we need to combine them

    # In case the headers come with empty columns between the columns with real values, the first ones are eliminated
    headers_list_1 = [column for column in df.columns if 'Unnamed' not in column]
    df = df[headers_list_1]

    # E.G. headers_list_1 = ['Unit', 'Unit Type', 'Unit.1', 'Resident', 'Name', 'Market', 'Charge', 'Amount', 'Resident.1', 'Other', 'Move In', 'Lease', 'Move Out', 'Balance', 'rent', 'petrent']
    headers_list_1 = df.columns.tolist()

    # we reformat the column names to unify them in all files
    for index,column in enumerate(headers_list_1):
        if '\n' in column:
            new_column = column.replace('\n',' ').replace('  ',' ')
            headers_list_1.remove(column)
            headers_list_1.insert(index,new_column)

    # If the df already has the headers in a single row, we return it untouched. Otherwise, we combine the two rows into one
    if ('Unit Sqft' in headers_list_1 or 'Charge Code' in headers_list_1):
        df.columns = headers_list_1
        return df
    else:
        # E.G. headers_list_2 = [nan, nan, 'Sq Ft', nan, nan, 'Rent', 'Code', nan, 'Deposit', 'Deposit', nan, 'Expiration', nan, nan, None, None]
        headers_list_2 = df.iloc[0].tolist()

        for index,element in enumerate(headers_list_2):
            if isinstance(element,str):
                if headers_list_1[index].endswith('.1'):
                    headers_list_1[index] = headers_list_1[index].replace('.1','') + " " + element
                else:
                    headers_list_1[index] = headers_list_1[index] + " " + element
        df.columns = headers_list_1
        return df

def extract_unique_values_from_charge_code(df):
    columns = df['Charge Code'].dropna().unique()
    return columns

def delete_unnecesary_rows(df):
    df = df.iloc[:-21]

    # since we merged rows 0 and 1 previously, row 0 now contains redundant information
    df =df.drop(0)

    # here we drop all the blank rows in the input sheet
    df.dropna(how="all",inplace=True)

    # We drop all the "totals" cells. E.g. in the workbook "Americana", in cell G9, you can see Total 958.
    for index,row in df.iterrows():
        if row['Charge Code'] == 'Total' or row['Resident'] == 'Total':
           df= df.drop(index)

    df = df.reset_index(drop=True)

    # We identify the cell where the section 'Future Residents/Applicants' starts so we can remove all the rows after that
    future_residents_index = df.index[df['Unit'] == 'Future Residents/Applicants']
    for index,row in df.iterrows():
        if index > future_residents_index:
            df= df.drop(index)

    # Remove 'Current Notice' and 'Future residents' rows
    row_to_delete = df[df['Unit'].notna() & df.iloc[:, 1:].isna().all(axis=1)]
    df = df.drop(row_to_delete.index)

    df = df.reset_index(drop=True)

    return df



def normalize_charge_code(df, column_name):
    df = df.reset_index(drop=True)
    # initialize columns with 0 so that we can sum the amount
    df[column_name]=0.0

    drop_list = []

    # calculate the correct value for each column E.G. 'rent', 'petrent', etc
    for index,row in df.iterrows():
        if index < len(df.index):
            if not pd.isna(row['Unit']):
                index_id = index
        if row['Charge Code'] == column_name:
            df.loc[index_id, column_name] += df.at[index, 'Amount']
            if index_id != index:
                drop_list.append(index)

    for row in drop_list:
        df.drop(row, inplace=True)
    df = df.reset_index(drop=True)

    return df


#---------------------------------------------------------------------------------------------

sheet_name = 'transformed_data'

# These two lines are for loading the name of the target workbook
load_dotenv()
file_name = os.getenv("file")

if not file_name:
    raise ValueError('Check that you entered your file name correctly in the .env file')

for server in servers:
    connection_string = f"mssql+pyodbc://{server['username']}:{server['password']}@{server['server_name']}/{server['database_name']}?driver=ODBC+Driver+17+for+SQL+Server"

    engine = create_engine(connection_string)
    df = read_file(file_name)

    df = build_headers(df)

    df = delete_unnecesary_rows(df)

    # Infer the new columns in the output sheet from charge code. E.G. 'rent', 'petrent', etc.
    columns = extract_unique_values_from_charge_code(df)

    for column in columns:
        df = normalize_charge_code(df, column)


    df['Lease Expiration'] = pd.to_datetime(df['Lease Expiration']).dt.strftime('%m/%d/%Y')
    df['Move Out'] = pd.to_datetime(df['Move Out']).dt.strftime('%m/%d/%Y')
    df['Move In'] = pd.to_datetime(df['Move In']).dt.strftime('%m/%d/%Y')
    df.rename(columns={'Charge Code': 'chargecode'}, inplace=True)
    #print(df.to_string())


    # Data for UW_RR Tables
    df.rename(columns={'Unit Type': 'UnitFloorPlan'}, inplace=True)
    df['SetAside'] = df['UnitFloorPlan'].astype(str).str[-2:].astype(int) / 100.0
    df["SetAside_Rounded"] = df['SetAside'].round(1)
    df["UnitType"] = "Residential"
    df["ResidentID"] = df["Resident"]
    df['OccupancyStatus'] = df['Resident'].apply(lambda x: 'Vacant' if x.lower() == 'vacant' else 'Occupied')
    df["Resident"] = df["Name"]
    with engine.connect() as connection:
        # Execute the SQL query
        deal_name = "Orchard Park"
        query = text(f"SELECT DealID FROM UW_Deals WHERE DealName='{deal_name}'")
        result = connection.execute(query)
        deal_id = result.scalar()
        df["DealID"] = deal_id

    df["AsOfDate"] = asofdate
    date = datetime.today().date()
    df["DateAdded"] = pd.to_datetime(date).strftime('%m/%d/%Y')
    df["Building"] = None
    with engine.connect() as connection:
        df_rr = pd.read_sql(f'''SELECT DISTINCT Unit, UnitFloorPlanID, Prospect_Unit_Type FROM UW_RR As RR
                         LEFT JOIN UW_RR_ChargeCodeAmount As CC ON CC.RRID = RR.RRID WHERE DealId = {deal_id} ''',connection)
        df["Unit"] = df["Unit"].str.strip()
        df = pd.merge(df, df_rr, on='Unit', how='inner')

    df.rename(columns={'Unit Sqft': 'Sq_Ft'}, inplace=True)
    df.rename(columns={'Market Rent': 'MarketRent'}, inplace=True)
    df.rename(columns={'Lease Expiration': 'LeaseExpiration'}, inplace=True)
    df.rename(columns={'Move In': 'MoveInDate'}, inplace=True)
    df.rename(columns={'Move Out': 'MoveOutDate'}, inplace=True)
    df["LeaseTerm"] = None
    df["UnitID"] = None
    df["Acquired"] = "Yes"
    df["LeaseStartDate"] = None
    df["LeaseType"] = None

    df_RR = df[['DealID', 'AsOfDate', 'DateAdded', 'Building','Unit', 'UnitType', 'UnitFloorPlan','UnitFloorPlanID', 'Prospect_Unit_Type','Sq_Ft',
                'ResidentID', 'Resident', 'LeaseType', 'OccupancyStatus', 'MarketRent','LeaseStartDate','LeaseExpiration','LeaseTerm',
                          'MoveInDate','MoveOutDate','SetAside','SetAside_Rounded','UnitID','Acquired']]
    #print(df_RR.to_string())

    try:
        df_RR.to_sql("UW_RR", engine, if_exists='append', index=False)
        print("Data successfully inserted into the SQL table.")
    except Exception as e:
        print(f"Error inserting data into SQL table: {e}")

    # with open(file_name, 'rb') as f:
    #     book = openpyxl.load_workbook(f)
    #
    #
    # if sheet_name in book.sheetnames:
    #     del book[sheet_name]
    # sheet = book.create_sheet(title=sheet_name)
    #
    # headers = df.columns.tolist()
    # sheet.append(headers)
    #
    # # Fill the new sheet with the data from the dataframe
    # for index, row in df.iterrows():
    #     row_values = row.tolist()
    #     sheet.append(row_values)
    #
    # with open(file_name, 'wb') as f:
    #     book.save(f)

    #Data for UW_RRChargecodeAmount
    with engine.connect() as connection:
       df_rrid = pd.read_sql(f"SELECT RRID, Unit FROM UW_RR WHERE [AsOfDate] = '{asofdate}'",connection)
    #print(df.to_string())

    df_char = pd.merge(df, df_rrid, on='Unit', how='inner')
    df_chargecode_amt = df_char[['chargecode','RRID','Amount']]
    condition = (df_chargecode_amt['Amount'] > 0)

    df_RR_chargecode_amt = df_chargecode_amt[condition]
    #print(df_RR_chargecode_amt.to_string())

    try:
        df_RR_chargecode_amt.to_sql("UW_RR_ChargeCodeAmount", engine, if_exists='append', index=False)
        print("Data successfully inserted into the SQL table.")
    except Exception as e:
        print(f"Error inserting data into SQL table: {e}")
