import dateutil
import pandas as pd
import os
import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
from et_xmlfile import xmlfile
from openpyxl import load_workbook
import numpy
import dateutil
import pytz
import six
import tzdata

#Set up the database connection details
server_name = 'q3solutions-dev.database.windows.net'
database_name = 'PropertyManager-dev'
username = 'q3solutions-dev'
password = 'Q3testMcQ'

#server_name = 'q3solutions.database.windows.net'
#database_name = 'PropertyManager'
#username = 'burtonguster'
#password = 'heardaboutplut0!'

# Set up the Excel file path
#file_path = r'C:\Users\rdsku\Aberdeen Capital\Aberdeen Capital - Documents\Underwriting Templates\GreenIQ\RentRoll.xlsx'
file_path = "C:/xampp/htdocs/RentInsite/Q3/orchardPark/ceva_americana-main/Rent_Roll_40501.xlsx"

sheet_name = 'RentRoll'

# Create the connection string
connection_string = f'mssql+pyodbc://{username}:{password}@{server_name}/{database_name}?driver=ODBC+Driver+17+for+SQL+Server'

# Create the SQLAlchemy engine
engine = create_engine(connection_string)


def read_file(file_name):
    df = pd.read_excel(file_name,
			sheet_name = 0)
    index = df[df.iloc[:, 0] == 'Unit'].index[0]
    df = pd.read_excel(file_name,
			sheet_name = 0,
            skiprows=index+1)
    return df

def build_headers(df):
    # Since the headers are split in two rows E.G. Unit + Sq Ft, we need to combine them

    # In case the headers come with empty columns between the columns with real values, the first ones are eliminated
    headers_list_1 = [column for column in df.columns if 'Unnamed' not in column]
    df = df[headers_list_1]

    # E.G. headers_list_1 = ['Unit', 'Unit Type', 'Unit.1', 'Resident', 'Name', 'Market', 'Charge', 'Amount', 'Resident.1', 'Other', 'Move In', 'Lease', 'Move Out', 'Balance', 'rent', 'petrent']
    headers_list_1 = df.columns.tolist()

    # we reformat the column names to unify them in all files
    for index,column in enumerate(headers_list_1):
        if '\n' in column:
            new_column = column.replace('\n',' ').replace('  ',' ')
            headers_list_1.remove(column)
            headers_list_1.insert(index,new_column)

    # If the df already has the headers in a single row, we return it untouched. Otherwise, we combine the two rows into one
    if ('Unit Sqft' in headers_list_1 or 'Charge Code' in headers_list_1):
        df.columns = headers_list_1
        return df
    else:
        # E.G. headers_list_2 = [nan, nan, 'Sq Ft', nan, nan, 'Rent', 'Code', nan, 'Deposit', 'Deposit', nan, 'Expiration', nan, nan, None, None]
        headers_list_2 = df.iloc[0].tolist()

        for index,element in enumerate(headers_list_2):
            if isinstance(element,str):
                if headers_list_1[index].endswith('.1'):
                    headers_list_1[index] = headers_list_1[index].replace('.1','') + " " + element
                else:
                    headers_list_1[index] = headers_list_1[index] + " " + element
        df.columns = headers_list_1
        return df

def extract_unique_values_from_charge_code(df):
    columns = df['Charge Code'].dropna().unique()
    return columns

def delete_unnecesary_rows(df):
    df = df.iloc[:-21]

    # since we merged rows 0 and 1 previously, row 0 now contains redundant information
    df =df.drop(0)

    # here we drop all the blank rows in the input sheet 
    df.dropna(how="all",inplace=True)
    
    # We drop all the "totals" cells. E.g. in the workbook "Americana", in cell G9, you can see Total 958.
    for index,row in df.iterrows():
        if row['Charge Code'] == 'Total' or row['Resident'] == 'Total':
           df= df.drop(index)
    
    df = df.reset_index(drop=True) 

    # We identify the cell where the section 'Future Residents/Applicants' starts so we can remove all the rows after that
    future_residents_index = df.index[df['Unit'] == 'Future Residents/Applicants']
    for index,row in df.iterrows():
        if index > future_residents_index:
            df= df.drop(index)

    # Remove 'Current Notice' and 'Future residents' rows
    row_to_delete = df[df['Unit'].notna() & df.iloc[:, 1:].isna().all(axis=1)]
    df = df.drop(row_to_delete.index)

    df = df.reset_index(drop=True)
    
    return df 



def normalize_charge_code(df, column_name):
    df = df.reset_index(drop=True)  
    # initialize columns with 0 so that we can sum the amount
    df[column_name]=0.0

    drop_list = []

    # calculate the correct value for each column E.G. 'rent', 'petrent', etc
    for index,row in df.iterrows():
        if index < len(df.index):
            if not pd.isna(row['Unit']):
                index_id = index
        if row['Charge Code'] == column_name:
            df.loc[index_id, column_name] += df.at[index, 'Amount']
            if index_id != index:
                drop_list.append(index)
    
    for row in drop_list:
        df.drop(row, inplace=True)
    df = df.reset_index(drop=True)

    return df


#---------------------------------------------------------------------------------------------

sheet_name = 'transformed_data'

# These two lines are for loading the name of the target workbook
load_dotenv()
file_name = os.getenv("file")

if not file_name:
    raise ValueError('Check that you entered your file name correctly in the .env file')
	
df = read_file(file_name)

df = build_headers(df)

df = delete_unnecesary_rows(df)

# Infer the new columns in the output sheet from charge code. E.G. 'rent', 'petrent', etc.
columns = extract_unique_values_from_charge_code(df)

for column in columns:
    df = normalize_charge_code(df, column)


df['Lease Expiration'] = pd.to_datetime(df['Lease Expiration']).dt.strftime('%m/%d/%Y')
df['Move Out'] = pd.to_datetime(df['Move Out']).dt.strftime('%m/%d/%Y')
df['Move In'] = pd.to_datetime(df['Move In']).dt.strftime('%m/%d/%Y')
df = df.drop(["Amount","Charge Code"],axis=1)

with open(file_name, 'rb') as f:
    book = openpyxl.load_workbook(f)


if sheet_name in book.sheetnames:
    del book[sheet_name]
sheet = book.create_sheet(title=sheet_name)

headers = df.columns.tolist()
sheet.append(headers)

# Fill the new sheet with the data from the dataframe
for index, row in df.iterrows():
    row_values = row.tolist()
    sheet.append(row_values)

with open(file_name, 'wb') as f:
    book.save(f)
